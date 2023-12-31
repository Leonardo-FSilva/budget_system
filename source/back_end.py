from datetime import datetime
import PySimpleGUI as sg
from pathlib import Path
import json
import pandas as pd
from openpyxl import load_workbook


DATA = Path(__file__).parent.parent / "config\data.json"
NEW_FILE_PAHT = Path(__file__).parent.parent / "config"
MODELO_EXCEL = Path(__file__).parent.parent / 'config\modelo_excel.xlsx'
RELATORIO_EXCEL = Path(__file__).parent.parent / 'config\\relatorio_excel.xlsx'

class User:
    def tela_editarlinha(self):
        return [[sg.Text('indice:'), sg.Input(size=7, key='editarlinha_ind'), sg.Text('descrição:'), sg.Input(size=80, key='editarlinha_des'), sg.Text('valor:'), sg.Input(size=15, key='editarlinha_val')],
                [sg.Button('Editar linha')]]
    def adcionaMaterial(self, desc, val):
        with open(DATA, 'r') as file:
            relatoriomp = json.load(file)

        relatoriomp["material"]["index"].append(len(relatoriomp["material"]['index'])+1)
        relatoriomp["material"]["descricao"].append(desc)
        relatoriomp["material"]["valor"].append(val)

        with open(DATA, 'w') as file:
            json.dump(relatoriomp, file, indent=4)
    def editarMaterial(self, ind, desc, val):
        with open(DATA, 'r') as file:
            relatoriomp = json.load(file)
        
        relatoriomp["material"]["descricao"][ind-1] = desc
        relatoriomp["material"]["valor"][ind-1] = val

        with open(DATA, 'w') as file:
            json.dump(relatoriomp, file, indent=4)
    def adicionaMaoObra(self, desc, val):
        with open(DATA, 'r') as file:
            relatoriomo = json.load(file)

        relatoriomo["maodeobra"]["index"].append(len(relatoriomo["maodeobra"]['index'])+1)
        relatoriomo["maodeobra"]["descricao"].append(desc)
        relatoriomo["maodeobra"]["valor"].append(val)

        with open(DATA, 'w') as file:
            json.dump(relatoriomo, file, indent=4)
    def editarMaoObra(self, ind, desc, val):
        with open(DATA, 'r') as file:
            relatoriomo =json.load(file)
        
        relatoriomo["maodeobra"]["descricao"][ind-1] = desc
        relatoriomo["maodeobra"]["valor"][ind-1] = val

        with open(DATA, 'w') as file:
            json.dump(relatoriomo, file, indent=4)
    def gerarRelatorio(self, path_to_open=''):
        if path_to_open == '':
            with open(DATA, 'r') as file:
                relatorio = json.load(file)
        else:
            with open(path_to_open, 'r')as file:
                relatorio = json.load(file)

        relatoriomp_formatado = ''
        relatoriomp_formatado += ('{:^7}| {:<100}|{}\n'.format('IND', 'Descrição', 'Valor R$'))
        relatoriomp_formatado += (f'{"-"*7}|{"-"*101}|{"-"*18}\n')
        if len(relatorio["material"]['index']) > 0:
            for iten in range(0, len(relatorio["material"]['index'])):
                relatoriomp_formatado += ('{:^7}| {:<100}| R${}\n'.format(relatorio["material"]['index'][iten], relatorio["material"]['descricao'][iten], relatorio["material"]['valor'][iten]))
        
        relatoriomo_formatado = ''
        relatoriomo_formatado += ('{:^7}| {:<100}|{}\n'.format('IND', 'Descrição', 'Valor R$'))
        relatoriomo_formatado += (f'{"-"*7}|{"-"*101}|{"-"*18}\n')
        if len(relatorio["maodeobra"]['index']) > 0:
            for iten in range(0, len(relatorio["maodeobra"]['index'])):
                relatoriomo_formatado += ('{:^7}| {:<100}| R${}\n'.format(relatorio["maodeobra"]['index'][iten], relatorio["maodeobra"]['descricao'][iten], relatorio["maodeobra"]['valor'][iten]))
        
        return [relatoriomo_formatado ,relatoriomp_formatado]
    def salvarRelatorio(self):
        with open(DATA, 'r') as file:
            relatorio = json.load(file)
 

        relatorio_nome = '\{} - {}.json'.format(str(datetime.now())[0:10], relatorio["informacoes"]["cliente"][0:3])
        relatorio_nome = str(NEW_FILE_PAHT) + relatorio_nome

        try:
            new_file = open(relatorio_nome, 'w')
            json.dump(relatorio, new_file, indent=4)
        except:
            return print("Error: avise o leonardo...")
        finally:
            new_file.close()
    def abrirRelatorio(self, path):
        with open(path, "r") as file:
            self.gerarRelatorio(file)
    def calcularRelatorio(self):
        with open(DATA, 'r') as file:
            relatorio = json.load(file)

        valor_maoobra = 0
        valor_mat = 0

        for i in range(len(relatorio["maodeobra"]["index"])):
            valor_maoobra += float(relatorio["maodeobra"]["valor"][i])

        

        for i in range(len(relatorio["material"]["index"])):
            valor_mat += float(relatorio["material"]["valor"][i])

        valor = valor_maoobra + valor_mat

        valor_maoobra = str('R${:.2f}'.format(valor_maoobra).replace('.',','))
        valor_mat = str('R${:.2f}'.format(valor_mat).replace('.',','))

        valor /= relatorio['coeficientes']['comissao']
        valor /= relatorio['coeficientes']['impostos']
        valor /= relatorio['coeficientes']['maodeobra']
        valor /= relatorio['coeficientes']['material']
        valor *= relatorio['coeficientes']['coeficientedeestrutura']
    
        valor = str('R${:.2f}'.format(valor)).replace('.',',')

        return([valor_mat, valor_maoobra, valor])
    def removerLinha(self, type, linha):
        with open(DATA, 'r') as file:
            relatorio = json.load(file)
        
        relatorio[type]['index'].pop()
        relatorio[type]['descricao'].pop(linha-1)
        relatorio[type]['valor'].pop(linha-1)

        with open(DATA, 'w') as file:
            json.dump(relatorio, file, indent=4)
    def alterarSetup(self, comissao, impostos, maodeobra, material, coeficientedeestrutura):
        with open(DATA, 'r') as file:
            relatorio = json.load(file)

        relatorio['coeficientes']['comissao'] = comissao
        relatorio['coeficientes']['impostos'] = impostos
        relatorio['coeficientes']['maodeobra'] = maodeobra
        relatorio['coeficientes']['material'] = material
        relatorio['coeficientes']['coeficientedeestrutura'] = coeficientedeestrutura

        with open(DATA, 'w') as file:
            json.dump(relatorio, file, indent=4)
    def gerarRelatorioFinal(self, cliente, estado, cidade, telefone, email, responsavel):
        with open(DATA, 'r') as file:
            relatorio = json.load(file)

        relatorio['informacoes']['cliente'] = cliente
        relatorio['informacoes']['estado'] = estado
        relatorio['informacoes']['cidade'] = cidade
        relatorio['informacoes']['telefone'] = telefone
        relatorio['informacoes']['email'] = email
        relatorio['informacoes']['responsavel'] = responsavel

        with open(DATA, 'w') as file:
            json.dump(relatorio, file, indent=4)

        # Criar o DataFrame a partir do dicionário
        df_maodeobra = pd.DataFrame(relatorio['maodeobra'])
        df_material = pd.DataFrame(relatorio['material'])

        book = load_workbook(MODELO_EXCEL)

        sheet = book.worksheets[0]

        start_row = 13
        start_column = 1

        for r in df_maodeobra.values:
            for i, value in enumerate(r):
                sheet.cell(row=start_row, column=start_column + i, value=value)
            start_row += 1

#___________________________________________________
        start_row = 13
        start_column = 5

        for r in df_material.values:
            for i, value in enumerate(r):
                sheet.cell(row=start_row, column=start_column + i, value=value)
            start_row += 1
#____________________________________________________
        celula = 'B4'  # cliente
        sheet[celula] = cliente

        celula = 'B5'  # estado
        sheet[celula] = estado

        celula = 'B6'  # cidade
        sheet[celula] = cidade

        celula = 'B7'  # telefone
        sheet[celula] = telefone

        celula = 'B8'  # email
        sheet[celula] = email

        celula = 'B9'  # responsavel
        sheet[celula] = responsavel
#___________________________________________________
        valores = self.calcularRelatorio()

        celula = 'F4'  # material
        sheet[celula] = valores[0]

        celula = 'F5'  # mao de obra
        sheet[celula] = valores[1]

        celula = 'F6'  # valor
        sheet[celula] = valores[2]

        
        book.save(RELATORIO_EXCEL)


if __name__ == "__main__":
    carlos = User()

    #carlos.adcionaMaterial('batata', 75)

    #carlos.adcionaMaterial('macarrao', 50)

    #carlos.adcionaMaterial('uva', 15)

    #carlos.editarMaterial(1, "maca do amor", 15)

    #relat=carlos.gerarRelatorio()
    #print(relat[0])

    #carlos.salvarRelatorio()
    #print(carlos.calcularRelatorio())